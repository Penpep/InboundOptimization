import streamlit as st
st.set_page_config(page_title="Inbound Dock Inventory Analysis", layout='wide',
    page_icon='smiley')


import pandas as pd
import math
from datetime import datetime, timedelta
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import io
import tempfile
from openpyxl.styles import PatternFill


def generate_times(start, end, cadence):
    if cadence == 0 or cadence is None:
        return []  # No deliveries
    interval = (end - start) / cadence
    return [start + i * interval for i in range(cadence)]

def generate_deliveries(qty, pack_size, cadence, shft_hrs, cons_rate, on_hand=0):
    if cadence == 0 or cadence is None:
        return []

    deliveries = []
    interval = shft_hrs / cadence
    total_needed_units = qty
    total_delivered_units = 0
    available_on_hand = on_hand

    for i in range(cadence):
        remaining_units = max(total_needed_units - total_delivered_units, 0)
        if remaining_units == 0:
            deliveries.append(0)
            continue

        interval_consumption = cons_rate * interval

        # How much is already on hand to cover consumption?
        if available_on_hand >= interval_consumption:
            available_on_hand -= interval_consumption
            deliveries.append(0)
            continue

        shortfall_units = interval_consumption - available_on_hand

        # Only deliver what remains needed
        deliver_units = min(shortfall_units, remaining_units)

        # Convert to packages (round up), but cap at remaining_units
        packages = math.ceil(deliver_units / pack_size)
        delivered_units = packages * pack_size

        if delivered_units > remaining_units:
            # Adjust down if over
            delivered_units = remaining_units
            packages = math.ceil(delivered_units / pack_size)

        total_delivered_units += delivered_units
        available_on_hand += delivered_units - interval_consumption

        deliveries.append(packages)

    return deliveries

def get_dock_inventory_peaks_per_part(deliveries, pack_size, consumption_rate, shift_hours, on_hand_dock, on_hand_lineside):
    if not deliveries or pack_size <= 0 or consumption_rate <= 0:
        return [0] * len(deliveries)

    interval = shift_hours / len(deliveries)
    dock_inventory_units = on_hand_dock
    lineside_inventory_units = on_hand_lineside

    dock_timeline = []

    for delivery in deliveries:
        delivered_units = delivery * pack_size
        dock_inventory_units += delivered_units
        dock_timeline.append(math.ceil(dock_inventory_units / pack_size))

        consumption = consumption_rate * interval
        pull_needed = max(consumption - lineside_inventory_units, 0)
        actual_pull = min(pull_needed, dock_inventory_units)

        dock_inventory_units -= actual_pull
        lineside_inventory_units += actual_pull
        lineside_inventory_units = max(0, lineside_inventory_units - consumption)

    return dock_timeline, lineside_inventory_units

# === Core Processing Logic ===

def run_analysis(uploaded_file, input_drive_unit):
    wb = load_workbook(uploaded_file, data_only=True)
    sheet_name = f'Inbound-{input_drive_unit}'
    ws = wb[sheet_name]

    # Determining lane material based on drive unit 
    if input_drive_unit == 'Hercules':
        side_lane = [
        "600-01361", "400-01256", "400-01259", "400-01260-C2",
        "600-01051", "600-01248", "600-02000", "600-02018", "600-00986"
        ] 

        lane = [
        "400-01318", "400-01950", "600-01020", "600-01035",
        "600-02306", "400-01226-C2", "400-01227-C2"
        ]


    # Read Parameters 
    cadence_shift_1 = ws['B5'].value
    cadence_shift_2 = ws['B13'].value
    lines_shift_1 = ws['B9'].value
    lines_shift_2 = ws['B10'].value
    box_dock_space = ws['D2'].value
    pallet_per_lane = ws['D3'].value
    side_lane_pallet = ws['D4'].value

    start_shift_1 = datetime.strptime("6:15", "%H:%M")
    end_shift_1 = datetime.strptime("15:00", "%H:%M")
    start_shift_2 = datetime.strptime("15:00", "%H:%M")
    end_shift_2 = datetime.strptime("23:15", "%H:%M")
    shift_1_hours = (end_shift_1 - start_shift_1).seconds / 3600
    shift_2_hours = (end_shift_2 - start_shift_2).seconds / 3600

    time_1 = generate_times(start_shift_1, end_shift_1, cadence_shift_1)
    time_2 = generate_times(start_shift_2, end_shift_2, cadence_shift_2)

    df_bom = pd.read_excel(uploaded_file, sheet_name=sheet_name, skiprows=15, header=None)
    df_bom.columns = [
        'Part Number', 'Description', 'Quantity / Unit', 'Needed per day',
        'Quantity Needed for Shift 1', 'Quantity Needed for Shift 2',
        'Pallets Utilized for Shift 1', 'Pallets Utilized for Shift 2',
        'Consumption Rate Units/ Hour Shift 1',
        'Consumption Rate / Hour Shift 2', 'Standard Pack Size', 'Package Type',
        'Maximum Storage on Lineside', 'Minimum Storage on Lineside', 
        'On-hand qty', 'QTY vs Shift 1', 'On-hand on dock'
    ]

    columns = ['Part Number', 'Package Type', 'Description']
    columns += [f"Delivery {i+1} (S1 - {t.strftime('%I:%M %p')})" for i, t in enumerate(time_1)]
    columns += [f"Delivery {i+1} (S2 - {t.strftime('%I:%M %p')})" for i, t in enumerate(time_2)]

    delivery_plan = []

    for _, row in df_bom.iterrows():
        part = row['Part Number']
        pkg_type = row['Package Type']
        pack_size = row['Standard Pack Size']
        descrip = row['Description']
        qty1 = row['Quantity Needed for Shift 1']
        qty2 = row['Quantity Needed for Shift 2']
        cons_1 = row['Consumption Rate Units/ Hour Shift 1']
        cons_2 = row['Consumption Rate / Hour Shift 2']
        on_hand = row['On-hand qty'] 

        net_qty1 = max(qty1 - on_hand, 0)
        remaining_on_hand = max(on_hand - qty1, 0)
        net_qty2 = max(qty2 - remaining_on_hand, 0)

        deliveries_1 = generate_deliveries(net_qty1, pack_size, cadence_shift_1, shift_1_hours, cons_1, on_hand)
        deliveries_2 = generate_deliveries(net_qty2, pack_size, cadence_shift_2, shift_2_hours, cons_2, remaining_on_hand)

        delivery_plan.append([part, pkg_type, descrip] + deliveries_1 + deliveries_2)

    df_output = pd.DataFrame(delivery_plan, columns=columns)

    # Caculate Total Number of Boxes and Pallets 
    df_boxes = df_output[df_output['Package Type'] == 'Box']
    df_pallets = df_output[df_output['Package Type'] == 'Pallet']

    box_total = ['TOTAL - BOX', 'Box', ''] + list(df_boxes.iloc[:, 3:].sum())
    pallet_total = ['TOTAL - PALLET', 'Pallet', ''] + list(df_pallets.iloc[:, 3:].sum())
    box_sums = df_boxes.iloc[:, 3:].sum()
    pallet_sums = df_pallets.iloc[:, 3:].sum()
    df_output.loc[len(df_output)] = pd.Series(dtype=object)
    df_output.loc[len(df_output)] = box_total
    df_output.loc[len(df_output)] = pallet_total

    # Calculate Rack Percent Usage 
    box_percent_usage = ['Box % Usage', 'Box', ''] + list(((box_sums / box_dock_space)*100).round(2))
    df_output.loc[len(df_output)] = pd.Series(dtype=object)
    df_output.loc[len(df_output)] = box_percent_usage

    # Calculate Side Pallet Space Utilization 
    side_df = df_output[df_output['Part Number'].isin(side_lane)]
    side_usage_total = ['TOTAL - SIDE LANE', '', ''] + list(side_df.iloc[:, 3:].sum())
    df_output.loc[len(df_output)] = side_usage_total

    side_pallet_usage = ['SIDE LANE % Usage', '', ''] + list(((side_df.iloc[:, 3:].sum() / side_lane_pallet) * 100).round(2))
    df_output.loc[len(df_output)] = side_pallet_usage
    df_output.loc[len(df_output)] = pd.Series(dtype=object)

    # Calculate lane percent Usage for each Lane material 
    lane_df = df_output[df_output['Part Number'].isin(lane)].copy()
    for idx, row in lane_df.iterrows():
        part = row['Part Number']
        pkg_type = row['Package Type']
        desc = row['Description']

        deliveries = row.iloc[3:]  
        usage_pct = (deliveries / pallet_per_lane) * 100

        label = f"LANE % - {part}"
        new_row = [label, pkg_type, desc] + list(usage_pct)
        df_output.loc[len(df_output)] = new_row
    
    # Dock Inventory Analysis
    space_records = []

    for idx, row in df_bom.iterrows():
        part = row['Part Number']
        pack_size = row['Standard Pack Size']
        pkg_type = row['Package Type']
        cons_1 = row['Consumption Rate Units/ Hour Shift 1'] * 0.9
        cons_2 = row['Consumption Rate / Hour Shift 2'] * 0.9
        on_hand_on_dock = row['On-hand on dock']
        on_hand = row['On-hand qty']
        remaining_on_hand = max(on_hand_on_dock - row['Quantity Needed for Shift 1'], 0)

        deliveries_1 = df_output.loc[idx, df_output.columns.str.contains(r"S1 -")].tolist()
        deliveries_2 = df_output.loc[idx, df_output.columns.str.contains(r"S2 -")].tolist()

        timeline_1, lineside2 = get_dock_inventory_peaks_per_part(deliveries_1, pack_size, cons_1, shift_1_hours, on_hand_on_dock, on_hand - on_hand_on_dock)
        timeline_2, _ = get_dock_inventory_peaks_per_part(deliveries_2, pack_size, cons_2, shift_2_hours, remaining_on_hand, lineside2)

        for i, inv in enumerate(timeline_1):
            space_records.append({'Part Number': part, 'Shift': 1, 'Delivery Label': f"Delivery {i+1} (S1 - {time_1[i].strftime('%I:%M %p')})", 'Inventory Packages': inv, 'Package Type': pkg_type})
        for i, inv in enumerate(timeline_2):
            space_records.append({'Part Number': part, 'Shift': 2, 'Delivery Label': f"Delivery {i+1} (S2 - {time_2[i].strftime('%I:%M %p')})", 'Inventory Packages': inv, 'Package Type': pkg_type})

    df_space = pd.DataFrame(space_records)

    # Pivot table to get final dock space report
    flat_records = []
    part_order = df_bom['Part Number'].tolist()

    for part in part_order:
        part_rows = df_space[df_space['Part Number'] == part]
        if part_rows.empty:
            continue
        pkg_type = part_rows['Package Type'].iloc[0]
        row = {'Part Number': part, 'Package Type': pkg_type, 'Description':df_bom[df_bom['Part Number'] == part]['Description'].iloc[0]}
        for _, rec in part_rows.iterrows():
            row[rec['Delivery Label']] = rec['Inventory Packages']
        flat_records.append(row)

    df_dock_space = pd.DataFrame(flat_records).fillna(0)


    df_boxes = df_dock_space[df_dock_space['Package Type'] == 'Box']
    df_pallets = df_dock_space[df_dock_space['Package Type'] == 'Pallet']

    box_total = ['TOTAL - BOX', 'Box', ''] + list(df_boxes.iloc[:, 3:].sum())
    pallet_total = ['TOTAL - PALLET', 'Pallet', ''] + list(df_pallets.iloc[:, 3:].sum())
    box_sums = df_boxes.iloc[:, 3:].sum()
    pallet_sums = df_pallets.iloc[:, 3:].sum()
    df_dock_space.loc[len(df_dock_space)] = pd.Series(dtype=object)
    df_dock_space.loc[len(df_dock_space)] = box_total
    df_dock_space.loc[len(df_dock_space)] = pallet_total

    # Calculate Rack Percent Usage 
    box_percent_usage = ['Box % Usage', 'Box', ''] + list(((box_sums / box_dock_space)*100).round(2))
    df_dock_space.loc[len(df_dock_space)] = pd.Series(dtype=object)
    df_dock_space.loc[len(df_dock_space)] = box_percent_usage

    # Calculate Side Pallet Space Utilization 
    side_df = df_dock_space[df_dock_space['Part Number'].isin(side_lane)]
    side_usage_total = ['TOTAL - SIDE LANE', '', ''] + list(side_df.iloc[:, 3:].sum())

    df_dock_space.loc[len(df_dock_space)] = side_usage_total

    side_pallet_usage = ['SIDE LANE % Usage', '', ''] + list(((side_df.iloc[:, 3:].sum() / side_lane_pallet) * 100).round(2))
    df_dock_space.loc[len(df_dock_space)] = side_pallet_usage
    df_dock_space.loc[len(df_dock_space)] = pd.Series(dtype=object)

    # Calculate lane percent Usage for each Lane material 
    lane_df = df_dock_space[df_dock_space['Part Number'].isin(lane)].copy()
    for idx, row in lane_df.iterrows():
        part = row['Part Number']
        pkg_type = row['Package Type']
        desc = row['Description']

        timeline = row.iloc[3:]  # Get delivery columns
        usage_pct = (timeline / pallet_per_lane) * 100

        label = f"LANE % - {part}"
        new_row = [label, pkg_type, desc] + list(usage_pct)
        df_dock_space.loc[len(df_dock_space)] = new_row

    return df_output, df_dock_space, side_lane, lane

# Color Key 
def add_color_key_to_side(ws):
    sand_fill = PatternFill(start_color="F4E1C1", end_color="F4E1C1", fill_type="solid")
    sand_blue_fill = PatternFill(start_color="B0C4DE", end_color="B0C4DE", fill_type="solid")
    sand_green_fill = PatternFill(start_color="C1D1C1", end_color="C1D1C1", fill_type="solid")
    sand_gray_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")

    legend_items = [
        ("Side Lane Part / Usage", sand_fill),
        ("Lane Part / Usage", sand_blue_fill),
        ("Totals", sand_green_fill),
        ("Racks", sand_gray_fill)
    ]

    # Find starting column (2 columns after last column of data)
    start_col = ws.max_column + 2
    start_row = 2  # You can change where vertically you want it

    for i, (label, fill) in enumerate(legend_items):
        cell_label = ws.cell(row=start_row + i, column=start_col, value=label)
        cell_color = ws.cell(row=start_row + i, column=start_col + 1)
        cell_color.fill = fill

#   Styling Helper Functions 
def highlight_side_lane_ws(ws, side_lane, lane):
    sand_fill = PatternFill(start_color="F4E1C1", end_color="F4E1C1", fill_type="solid")
    sand_blue_fill = PatternFill(start_color="B0C4DE", end_color="B0C4DE", fill_type="solid")
    sand_green_fill = PatternFill(start_color="C1D1C1", end_color="C1D1C1", fill_type="solid")
    sand_gray_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")

    header = [cell.value for cell in ws[1]]
    part_num_col_idx = header.index("Part Number") + 1
    pkg_type_col_idx = header.index("Package Type") + 1

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        part_value = str(row[part_num_col_idx - 1].value).strip()
        pkg_type_value = str(row[pkg_type_col_idx - 1].value).strip()

        if part_value in side_lane or "SIDE LANE % Usage" in part_value or "TOTAL - SIDE LANE" in part_value:
            fill = sand_fill
        elif part_value in lane or part_value.startswith("LANE % -"):
            fill = sand_blue_fill
        elif "TOTAL - BOX" in part_value or "TOTAL - PALLET" in part_value:
            fill = sand_green_fill
        elif pkg_type_value == "Box":
            fill = sand_gray_fill
        else:
            fill = None

        if fill:
            for cell in row:
                cell.fill = fill

    add_color_key_to_side(ws)
    
   # Streamlit interface
st.title("Inbound Delivery Planning Tool")

valid_names = ['Proteus', 'Hercules', 'Megasus'] 
st.write("Make sure you have sheets titled 'Inbound-Proteus', 'Inbound-Hercules', or 'Inbound-Megasus' in your Excel file.")
input_drive_unit = st.selectbox("Select the Drive Unit", valid_names)

uploaded_file = st.file_uploader("Upload your BOM Excel file", type=["xlsx", "xlsm"])
if uploaded_file:
    st.success("File uploaded successfully.")
    if st.button("Generate Delivery Plan"):
        with st.spinner("Processing..."):
             df_output, df_dock_space, side_lane, lane = run_analysis(uploaded_file, input_drive_unit)


        st.subheader(f"Delivery Plan Output for {input_drive_unit}")

        st.dataframe(df_output)
        st.subheader("Dock Inventory Space Per Part (Pallet Equivalents)")
        st.dataframe(df_dock_space)

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_output.to_excel(writer, sheet_name="Delivery Plan", index=False)
            df_dock_space.to_excel(writer, sheet_name="Dock Inventory Space", index=False)
            workbook = writer.book
            ws1 = workbook["Delivery Plan"]
            ws2 = workbook["Dock Inventory Space"]

            # Highlight in-memory worksheets
            highlight_side_lane_ws(ws1, side_lane, lane)
            highlight_side_lane_ws(ws2, side_lane, lane)

        output.seek(0)
        st.download_button(
            "Download Excel",
            output,
            file_name=f"Delivery_Plan_{input_drive_unit}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        st.write("Download Excel File for data to be color coded")
